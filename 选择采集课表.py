xnxq01id_list = [
    "2025-2026-2", "2025-2026-1", "2024-2025-2", "2024-2025-1", "2023-2024-2", "2023-2024-1", "2022-2023-2", "2022-2023-1",
]
kbjcmsid_list = [
    ("04185F9CDDC04BC2AF96C38D2B31EB68", "长江新区校区作息时间"),
    ("952E3FC5FF563C09E053459072CA5D79", "武昌校区作息时间")
]
import requests
import openpyxl
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup
import csv
import webbrowser
from app_paths import resolve_input_path, resolve_output_path

def choose_from_list(options, label, key='name'):
    print(f"请选择{label}：")
    for idx, item in enumerate(options):
        print(f"{idx+1}. {item[key] if isinstance(item, dict) else item}")
    while True:
        try:
            sel = int(input(f"输入序号(1-{len(options)}): "))
            if 1 <= sel <= len(options):
                return options[sel-1]
        except Exception:
            pass
        print("输入有误，请重新输入。")


def get_jsessionid():
    webbrowser.open("https://jwxt.hubu.edu.cn/jsxsd/framework/xsMain.jsp")
    print("已为你打开教务系统页面。登录后进入课表查询界面，按F12打开开发者工具，切换到Application/存储/Storage->Cookies，找到JSESSIONID（路径为根目录），复制其值并粘贴到下方：")
    return input("请输入JSESSIONID: ").strip()

BASE_URL = "https://jwxt.hubu.edu.cn"
KB_URL = BASE_URL + "/kkglAction.do?method=toKbforward"

# 自动批量采集课表，参数来自csv
def get_kb_html(session, xnxq01id, kbjcmsid, yxbh, rxnf, zy, bjbh, xx04mc):
    data = {
        'method': 'toKbforward',
        'type': 'xx04',
        'isview': '1',
        'zc': '',
        'xnxq01id': xnxq01id,
        'kbjcmsid': kbjcmsid,
        'yxbh': yxbh,
        'rxnf': rxnf,
        'zy': zy,
        'bjbh': bjbh,
        'xx04id': bjbh,
        'xx04mc': xx04mc,
    }
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36 Edg/139.0.0.0',
        'Referer': KB_URL,
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    }
    resp = session.post(KB_URL, headers=headers, data=data)
    return resp.text


def _looks_like_teacher(line: str) -> bool:
    if 1 <= len(line) <= 4 and all("\u4e00" <= ch <= "\u9fff" for ch in line):
        return True
    return False


def _normalize_course_text(raw_text: str) -> str:
    lines = [line.strip() for line in raw_text.splitlines()]
    blocks = []
    current = []
    for line in lines:
        if not line:
            continue
        if "----" in line:
            if current:
                blocks.append(current)
                current = []
            continue
        current.append(line)
    if current:
        blocks.append(current)

    cleaned_blocks = []
    for block in blocks:
        course_name = ""
        teacher = ""
        classroom = ""
        hours = ""
        exam_type = ""
        for line in block:
            if not line:
                continue
            if line in {"★", "●"}:
                continue
            if line.startswith("【") and line.endswith("】"):
                continue
            if "节" in line and "周" in line:
                continue
            if "班" in line:
                continue
            if line in {"考试", "考查"}:
                exam_type = line
                continue
            if "总学时" in line:
                hours = line.replace("总学时：", "学时:").replace("总学时:", "学时:")
                continue
            if line.startswith("(") and line.endswith(")") and any(k in line for k in ["讲课", "实验", "实践"]):
                if not hours:
                    hours = "学时:" + line.strip("()")
                continue
            if not course_name and "教室" not in line and "实验室" not in line and not line.isalnum():
                course_name = line
                continue
            if not teacher and _looks_like_teacher(line):
                teacher = line
                continue
            if not classroom and ("教室" in line or "实验室" in line) and not line.startswith("【"):
                classroom = line
                continue

        parts = [course_name, hours, classroom, teacher, exam_type]
        cleaned = " / ".join([p for p in parts if p])
        if cleaned:
            cleaned_blocks.append(cleaned)

    return "\n".join(cleaned_blocks)


def extract_full_cell(td):
    """提取一个td中所有kbcontent/kbcontent1的全部内容（包括隐藏的），保留多段课程"""
    cell_texts = []
    divs = td.find_all(lambda tag: tag.name == "div" and tag.get("class", [""])[0].startswith("kbcontent"))
    for div in divs:
        raw_txt = div.get_text(separator="\n", strip=True)
        if raw_txt:
            normalized = _normalize_course_text(raw_txt)
            if normalized:
                cell_texts.append(normalized)
    return "\n".join(cell_texts) if cell_texts else ""

def parse_kb_to_matrix(html):
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table", {"id": "kbtable"})
    if table is None:
        # 识别常见未登录/无数据页面
        text = soup.get_text(separator="", strip=True)
        if any(k in text for k in ["登录", "统一身份认证", "用户名", "密码", "验证码"]):
            raise ValueError("未找到课表table：很可能未登录或会话已失效，请重新获取JSESSIONID")
        if any(k in text for k in ["暂无数据", "没有课表", "未查询到"]):
            raise ValueError("未找到课表table：该班级在所选学期/校区下无课表")
        # 保存调试HTML，便于排查
        try:
            with open("kb_debug.html", "w", encoding="utf-8") as f:
                f.write(html)
        except Exception:
            pass
        raise ValueError("未找到课表table，已保存调试页面到 kb_debug.html")
    ths = table.find_all("tr")[0].find_all("th")
    days = [th.get_text(strip=True) for th in ths[1:]]
    matrix = []
    trs = table.find_all("tr")[1:-1]
    for tr in trs:
        tds = tr.find_all("td")
        if not tds: continue
        section = tds[0].get_text(separator="", strip=True)
        row = [section]
        for td in tds[1:]:
            cell = extract_full_cell(td)
            row.append(cell)
        matrix.append(row)
    return days, matrix

def export_matrix_to_excel(days, matrix, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "课表"
    ws.append(["节次"] + days)
    for row in matrix:
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 50))
    # 自动调整行高
    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value:
                lines = str(cell.value).count("\n") + 1
                max_lines = max(max_lines, lines)
        ws.row_dimensions[cell.row].height = max(15, min(max_lines * 15, 120))
    wb.save(filename)


def main():
    jsessionid = get_jsessionid()
    session = requests.Session()
    session.cookies.update({'JSESSIONID': jsessionid})
    # 读取参数csv
    input_csv = resolve_input_path("筛选的课表参数.csv")
    with open(input_csv, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = []
        for r in reader:
            # 统一去除key的BOM和空白
            row = {k.strip().replace('\ufeff',''): v for k, v in r.items()}
            rows.append(row)
    # 院系
    yx_list = []
    yx_seen = set()
    for row in rows:
        if row['yxbh'] and row['院系'] and (row['yxbh'], row['院系']) not in yx_seen:
            yx_list.append({'yxbh': row['yxbh'], 'name': row['院系']})
            yx_seen.add((row['yxbh'], row['院系']))
    yx = choose_from_list(yx_list, '院系')
    # 年级
    nj_list = []
    nj_seen = set()
    for row in rows:
        if row['yxbh'] == yx['yxbh'] and row['rxnf'] and (row['rxnf'], row['年级']) not in nj_seen:
            nj_list.append({'rxnf': row['rxnf'], 'name': row['年级']})
            nj_seen.add((row['rxnf'], row['年级']))
    nj = choose_from_list(nj_list, '年级')
    # 专业
    zy_list = []
    zy_seen = set()
    for row in rows:
        if row['yxbh'] == yx['yxbh'] and row['rxnf'] == nj['rxnf'] and row['zy'] and (row['zy'], row['专业']) not in zy_seen:
            zy_list.append({'zy': row['zy'], 'name': row['专业']})
            zy_seen.add((row['zy'], row['专业']))
    zy = choose_from_list(zy_list, '专业')
    # 班级
    bj_list = []
    for row in rows:
        if row['yxbh'] == yx['yxbh'] and row['rxnf'] == nj['rxnf'] and row['zy'] == zy['zy']:
            bj_list.append({'bjbh': row['bjbh'], 'name': row['班级'], 'xnxq01id': row.get('xnxq01id',''), 'kbjcmsid': row.get('kbjcmsid','')})
    bj = choose_from_list(bj_list, '班级')
    # 学年学期（固定列表选择）
    xnxq = choose_from_list([{'name': x} for x in xnxq01id_list], '学年学期')
    # 校区（固定列表选择）
    kbjc = choose_from_list([{'name': x[0], 'desc': x[1]} for x in kbjcmsid_list], '校区（04开头为长江新区校区，95开头为武昌校区）')
    # 找到最终参数行（只用bjbh匹配）
    param_row = None
    for row in rows:
        if row['bjbh'] == bj['bjbh']:
            param_row = row
            break
    if not param_row:
        print('未找到对应参数，退出')
        return
    # 采集课表
    try:
        html = get_kb_html(session, xnxq['name'], kbjc['name'], param_row['yxbh'], param_row['rxnf'], param_row['zy'], param_row['bjbh'], param_row['班级'])
        days, matrix = parse_kb_to_matrix(html)
        fname = f"{param_row['班级']}_{param_row['rxnf']}_{xnxq['name']}.xlsx".replace("/", "_").replace("[", "").replace("]", "").replace(" ", "")
        output_path = resolve_output_path(fname)
        export_matrix_to_excel(days, matrix, output_path)
        print(f"已保存: {output_path}")
    except Exception as e:
        print(f"采集失败: {e}")

if __name__ == "__main__":
    main()
